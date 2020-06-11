import csv
import os
import random
import re
import tempfile
import time
import traceback
from io import open
from urllib.parse import urlencode

import PyPDF2
import google
import requests
from PyPDF2.utils import PdfReadError
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from openpyxl import load_workbook
from pdf2image import convert_from_path
from pymongo import MongoClient
from pytesseract import image_to_string
from requests.exceptions import HTTPError
from urllib3.exceptions import InsecureRequestWarning

# client = MongoClient('localhost', 27017)
client = MongoClient(
    "mongodb://dbIntelligentMindInformation:MoRLaRGAbLevoChutL@intelligent-mind-information-pnrix.mongodb.net/mining-name?retryWrites=true&w=majority")
db = client.miningNames
tablesNames = db.names
tablesConfiguration = db.configuration


def telegram_bot_sendtext(bot_message):
    bot_token = '700094640:AAEKqh3z7e6LfiXTXeFTpRnugSioqMkVA7o'
    bot_chat_id = ('1048049017', '680337670')
    for chat_id in bot_chat_id:
        query = {
            "chat_id": chat_id,
            "parse_mode": "markdown",
            "text": bot_message
        }
        send_text = "https://api.telegram.org/bot{token}/sendMessage?{query}" \
            .format(token=bot_token, query=urlencode(query))
        respose = requests.get(send_text)
        print(respose.json())


def telegram_bot_delete_conversation():
    chat = requests.get("https://api.telegram.org/bot700094640:AAEKqh3z7e6LfiXTXeFTpRnugSioqMkVA7o/getUpdates")
    result_chat = chat.json()
    for conversation in result_chat.get('result'):
        message_id = conversation.get('message').get('message_id')
        chat_id = conversation.get('message').get('reply_to_message').get('chat').get('id')
        is_bot = conversation.get('message').get('reply_to_message').get('chat').get('is_bot')
        if is_bot:
            response_delete = requests.get(
                "https://api.telegram.org/bot700094640:AAEKqh3z7e6LfiXTXeFTpRnugSioqMkVA7o/deleteMessage" +
                "?chat_id=" + str(chat_id) + "&message_id=" + str(message_id)
            )
            print(response_delete.json())


def log():
    telegram_bot_sendtext(
        str(i + 1) + "\n" +
        "*" + name + "*\n" +
        "[link](" + link + ")\n" +
        "```log" + traceback.format_exc() + "```"
    )
    print(str([i + 1, name, link]))
    print(traceback.format_exc())


def read_pdf(response):
    text = ""
    try:
        name_and_directory_file = "/tmp/" + str(random.getrandbits(128)) + ".pdf"
        with open(name_and_directory_file, 'wb') as f:
            f.write(response.content)
            f.close()

        pdf = open(name_and_directory_file, "rb")
        read_pdf_extractor = PyPDF2.PdfFileReader(pdf, strict=False)
        max_pages_pdf = read_pdf_extractor.getNumPages()
        count_pages = 0
        while count_pages < max_pages_pdf:
            text += read_pdf_extractor.getPage(count_pages).extractText()
            count_pages += 1

        pdf.close()
        del pdf
        del read_pdf_extractor

        with tempfile.TemporaryDirectory() as path:
            for page in range(1, max_pages_pdf, 10):
                images_from_path = convert_from_path(
                    name_and_directory_file, dpi=200, first_page=page,
                    last_page=min(page + 10 - 1, max_pages_pdf), output_folder=path
                )

                for img in images_from_path:
                    text += image_to_string(img, lang='por')
        del images_from_path
        os.remove(name_and_directory_file)
    except PdfReadError:
        log()

    return text


clear = lambda: os.system('clear')
clear()

if os.path.exists("finding-cpf.csv"):
    os.remove("finding-cpf.csv")

file = open("finding-cpf.csv", "w+")
file.close()

workbook = load_workbook('/files-imports/lista-controle-precatorios.xlsx')
sheet = workbook.active
row_count = sheet.max_row

# read
with open('finding-cpf.csv', 'rb') as csvfile:
    spamreader = csv.reader(csvfile, delimiter=',', quotechar='|')
    for row in spamreader:
        print(', '.join(row))

headers = {
    'Accept-Encoding': 'gzip, deflate, sdch',
    'Accept-Language': 'en-US,en;q=0.8',
    'Upgrade-Insecure-Requests': '1',
    'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.87 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Cache-Control': 'max-age=0',
    'Connection': 'keep-alive',
}

findOne = tablesConfiguration.find_one()
count = findOne['rowFromCSV']
print(count)

ua = UserAgent(use_cache_server=False, verify_ssl=False)
with open('finding-cpf.csv', 'a') as csvfile:
    spamwriter = csv.writer(csvfile, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)

    for i in range(row_count):
        if i != 0 and i >= count - 1:
            name = sheet.cell(row=i + 1, column=2).value

            tablesConfiguration.update_one({"_id": findOne["_id"]}, {"$set": {"rowFromCSV": count}})

            search_results = google.search('"vestibular" AND "' + name + '"', 1)

            for result in search_results:
                try:
                    link = result.link
                    print(i + 1, name, link)

                    if tablesNames.find({"link": link}).count() == 0:
                        try:
                            response = requests.get(result.link, headers=headers)
                        except ConnectionError as e:
                            log()
                            break
                        except requests.exceptions.SSLError as e:
                            requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
                            response = requests.get(result.link, verify=False, headers=headers)

                        if response.status_code == 404:
                            break

                        if response.headers._store['content-type'][1].count('pdf'):
                            htmlText = read_pdf(response)
                        else:
                            html = BeautifulSoup(response.text, "html.parser")
                            htmlText = html.text
                            pass

                        # adicionar verificação de nome de advogado na string
                        findName = htmlText.count(name)
                        if findName:

                            findPhone = re.findall(r"(\(\d{2}\)\s?\d{4,5}-?\d{4})", htmlText)
                            findCPF = re.findall(r"(\d{3}\.\d{3}\.\d{3}-\d{2})", htmlText)

                            if len(findPhone) or len(findCPF):
                                list_cpfs = []
                                for values_cpfs in findCPF:
                                    list_cpfs.append({
                                        "cpf": values_cpfs,
                                        "name": ""
                                    })

                                tablesNames.insert_one({
                                    "name": name,
                                    "phone": findPhone,
                                    "cpf": list_cpfs,
                                    "link": link,
                                })
                                print(i + 1, name, link, str(findPhone), str(findCPF))
                                pass
                        pass

                    time.sleep(random.randint(1, 30))

                    count = i + 1
                except Exception:
                    log()
                    break
