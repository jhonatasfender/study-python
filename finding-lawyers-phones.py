import csv
import os
from google import google
import time
import random
from openpyxl import load_workbook
import requests
import re
from bs4 import BeautifulSoup
import os

clear = lambda: os.system('clear')
clear()

if os.path.exists("finding-lawyers-phones.csv"):
    os.remove("finding-lawyers-phones.csv")

file = open("finding-lawyers-phones.csv", "w+")
file.close()

workbook = load_workbook(
    '/home/joantas/Downloads/lista-controle-precatorios.xlsx')
sheet = workbook.active
row_count = sheet.max_row

# read
with open('finding-lawyers-phones.csv', 'rb') as csvfile:
    spamreader = csv.reader(csvfile, delimiter=',', quotechar='|')
    for row in spamreader:
        print(', '.join(row))

# append
with open('finding-lawyers-phones.csv', 'a') as csvfile:
    spamwriter = csv.writer(csvfile,
                            delimiter=',',
                            quotechar='|',
                            quoting=csv.QUOTE_MINIMAL)

    for i in range(row_count):
        if i != 0:
            name = sheet.cell(row=i + 1, column=3).value

            if name != None:
                search_results = google.search('"' + name + '" AND "telefone"',
                                               1)

                for link in search_results:
                    try:
                        response = requests.get(link.link)

                        html = BeautifulSoup(response.text, "html.parser")

                        find = re.findall(r"(\(\d{2}\)\s?\d{4}-?\d{4})",
                                          html.text)

                        for phone in find:
                            spamwriter.writerow([name, phone, link.link])

                            print(i + 1, [name, phone, link.link])

                    except Exception as err:
                        spamwriter.writerow([name, str(err), link.link])

                        print([name, str(err), link.link])

                time.sleep(random.randint(1, 30))

                pass

            pass
