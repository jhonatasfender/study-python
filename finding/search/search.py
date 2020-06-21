import os

from openpyxl import load_workbook

from finding.exception import SearchException
from finding.search.connection import Connection
from finding.search.processing_search import ProcessingSearch


class Search:
    __configuration = None

    def __init__(self, args):
        self.args = args
        self.clear_console()
        self.connection = Connection(args)

        if not self.args.name_term and not self.args.name_search:
            self.assign_current_reading_value_file()
            self.__get_csv()
            self.read_csv()
        elif self.args.name_term and self.args.name_search:
            self.processing_name()
        else:
            SearchException.choose_a_file_mining_type_or_simple_name()

    def processing_name(self):
        ProcessingSearch.start_single_name(
            self.connection, self.__configuration,
            self.args.name_search, self.args.name_term
        )

    def read_csv(self):
        for i in range(self.row_count):
            ProcessingSearch.start_with_csv(
                self.connection, i, self.sheet, self.__configuration,
                self.__configuration.get('rowFromCSV')
            )

    def assign_current_reading_value_file(self):
        self.__configuration = self.connection.get_configuration().find_one()

    def __get_csv(self):
        if self.args.file_in:
            self.workbook = load_workbook(self.args.file_in)
            self.sheet = self.workbook.active
            self.row_count = self.sheet.max_row
        else:
            raise ValueError('É obrigatório preencher o --file-in')

    @staticmethod
    def clear_console():
        os.system('clear')
